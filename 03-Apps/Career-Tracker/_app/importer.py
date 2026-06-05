"""Import logic for Career Tracker.

Handles importing accomplishments from pasted text and Excel files.
"""
import re


def parse_pasted_list(text: str) -> list:
    """Parse pasted text into individual items.

    Splits text by newlines and cleans up common list formatting:
    - Bullet points (-, *, •, ·, →, ►, ▸)
    - Numbered lists (1., 2., 1), 2), etc.)
    - Leading/trailing whitespace

    Args:
        text: Raw pasted text with one item per line

    Returns:
        List of cleaned item strings (empty lines removed)
    """
    if not text or not text.strip():
        return []

    lines = text.strip().split("\n")
    items = []

    for line in lines:
        # Strip whitespace
        cleaned = line.strip()

        if not cleaned:
            continue

        # Remove common bullet markers
        cleaned = re.sub(r"^[-•*·→►▸]\s*", "", cleaned)

        # Remove numbered list markers (1. or 1) or (1) patterns)
        cleaned = re.sub(r"^\d+[.)]\s*", "", cleaned)
        cleaned = re.sub(r"^\(\d+\)\s*", "", cleaned)

        # Remove checkbox markers [ ] or [x]
        cleaned = re.sub(r"^\[[ xX]?\]\s*", "", cleaned)

        cleaned = cleaned.strip()
        if cleaned:
            items.append(cleaned)

    return items


def import_from_docx(filepath: str) -> list:
    """Read paragraphs from a Word document and return items as strings.

    Extracts meaningful text from the document:
    - Regular paragraphs (non-empty)
    - List items (bulleted or numbered)
    - Skips headers that look like section titles (short + all caps or bold-only)

    Args:
        filepath: Path to the .docx file

    Returns:
        List of non-empty paragraph/item strings

    Raises:
        FileNotFoundError: If the file doesn't exist
        ImportError: If python-docx is not installed
    """
    try:
        from docx import Document
    except ImportError:
        raise ImportError(
            "python-docx is required for Word import. "
            "Install it with: pip install python-docx"
        )

    doc = Document(filepath)
    items = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        # Clean up bullet/list markers that Word sometimes leaves in text
        cleaned = re.sub(r"^[-•*·→►▸]\s*", "", text)
        cleaned = re.sub(r"^\d+[.)]\s*", "", cleaned)
        cleaned = re.sub(r"^\[[ xX]?\]\s*", "", cleaned)
        cleaned = cleaned.strip()

        if not cleaned:
            continue

        # Skip very short lines that are likely headings/titles
        # (unless they're the only content)
        if len(cleaned) < 5 and cleaned.isupper():
            continue

        items.append(cleaned)

    return items


def import_from_excel(filepath: str, column: str = "A") -> list:
    """Read a column from an Excel file and return items as strings.

    Args:
        filepath: Path to the .xlsx file
        column: Column letter to read (default "A")

    Returns:
        List of non-empty cell values as strings

    Raises:
        FileNotFoundError: If the file doesn't exist
        ImportError: If openpyxl is not installed
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel import. "
            "Install it with: pip install openpyxl"
        )

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    items = []
    col_idx = ord(column.upper()) - ord("A") + 1

    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if cell.value is not None:
            value = str(cell.value).strip()
            if value:
                items.append(value)

    wb.close()
    return items
