"""Configuration for Career Tracker."""
import json
from pathlib import Path

APP_TITLE = "Career Tracker"
APP_SIZE = "900x650"

# Paths
APP_DIR = Path(__file__).parent.resolve()
PROJECT_DIR = APP_DIR.parent
CONFIG_DIR = PROJECT_DIR / "config"
SETTINGS_FILE = CONFIG_DIR / "settings.json"
ENTRIES_FILE = CONFIG_DIR / "entries.json"

# Theme colors
DARK_THEME = {
    "bg": "#1e1e2e",
    "fg": "#cdd6f4",
    "accent": "#89b4fa",
    "surface": "#313244",
    "border": "#45475a",
    "muted": "#a6adc8",
    "green": "#a6e3a1",
    "red": "#f38ba8",
    "yellow": "#f9e2af",
}

LIGHT_THEME = {
    "bg": "#ffffff",
    "fg": "#1e1e2e",
    "accent": "#1a73e8",
    "surface": "#f5f5f5",
    "border": "#e0e0e0",
    "muted": "#666666",
    "green": "#2e7d32",
    "red": "#c62828",
    "yellow": "#f57f17",
}

MONTH_NAMES = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

STATUS_OPTIONS = [
    "Completed",
    "In Progress",
    "Launched",
    "Early Development",
    "Ongoing",
]

# STAR placeholders
PLACEHOLDER_SITUATION = "[What problem or gap existed?]"
PLACEHOLDER_TASK = "[What were you responsible for delivering?]"
PLACEHOLDER_ACTION = "[What specific actions did you take?]"
PLACEHOLDER_RESULT = "[What was the measurable outcome or impact?]"

# Supported file types for open/save
SUPPORTED_FILE_TYPES = [
    ("Career Tracker files", "*.json"),
    ("Word documents", "*.docx"),
    ("Excel files", "*.xlsx"),
    ("Markdown files", "*.md"),
    ("Text files", "*.txt"),
    ("All files", "*.*"),
]


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

def load_settings() -> dict:
    """Load user settings from settings.json."""
    if SETTINGS_FILE.exists():
        try:
            return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return {"theme": "dark", "default_export": "docx"}


def save_settings(settings: dict):
    """Save user settings to settings.json."""
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    SETTINGS_FILE.write_text(json.dumps(settings, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# Entries (workspace-aware)
# ---------------------------------------------------------------------------

# Active workspace file — when set, entries load/save from this file
_active_file: Path | None = None


def get_active_file() -> Path | None:
    """Get the currently active workspace file."""
    global _active_file
    if _active_file is None:
        # Check settings for last-opened file
        settings = load_settings()
        last_file = settings.get("last_opened_file")
        if last_file:
            p = Path(last_file)
            if p.exists():
                _active_file = p
    return _active_file


def set_active_file(filepath: Path | None):
    """Set the active workspace file. None means use default entries.json."""
    global _active_file
    _active_file = filepath

    # Persist to settings
    settings = load_settings()
    settings["last_opened_file"] = str(filepath) if filepath else None
    save_settings(settings)


def get_active_file_display() -> str:
    """Get a display-friendly name for the active file."""
    active = get_active_file()
    if active:
        return f"{active.stem} ({active.parent.name}/{active.name})"
    return "Default (config/entries.json)"


def load_entries() -> list:
    """Load career entries from the active file.

    If an active file is set, loads from that file (parsing based on extension).
    Otherwise, loads from the default entries.json.
    """
    active = get_active_file()

    if active and active.exists():
        return _load_from_file(active)

    # Default: load from entries.json
    if ENTRIES_FILE.exists():
        try:
            return json.loads(ENTRIES_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return []


def save_entries(entries: list):
    """Save career entries to the active file.

    If an active file is set, saves to that file (in its native format).
    Otherwise, saves to the default entries.json.
    """
    active = get_active_file()

    if active:
        _save_to_file(entries, active)
    else:
        # Default: save to entries.json
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        ENTRIES_FILE.write_text(
            json.dumps(entries, indent=2, ensure_ascii=False), encoding="utf-8"
        )


def create_new_file(filepath: Path) -> bool:
    """Create a new empty career tracker file and set it as active.

    Returns True if successful.
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    suffix = filepath.suffix.lower()
    if suffix == ".json":
        filepath.write_text("[]", encoding="utf-8")
    elif suffix == ".md":
        filepath.write_text(
            "# Career Tracker\n\n*Entries tracked in STAR format.*\n\n---\n\n",
            encoding="utf-8",
        )
    elif suffix == ".txt":
        filepath.write_text("", encoding="utf-8")
    elif suffix == ".docx":
        try:
            from docx import Document
            doc = Document()
            doc.add_heading("Career Tracker", level=1)
            doc.save(str(filepath))
        except ImportError:
            filepath.write_text("", encoding="utf-8")
            return False
    elif suffix == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Entries"
            ws.append(["Title", "Status", "Date", "Situation", "Task", "Actions", "Result"])
            wb.save(filepath)
        except ImportError:
            return False
    else:
        filepath.write_text("[]", encoding="utf-8")

    set_active_file(filepath)
    return True


# ---------------------------------------------------------------------------
# File I/O (format-specific loading and saving)
# ---------------------------------------------------------------------------

def _load_from_file(filepath: Path) -> list:
    """Load entries from a file based on its extension."""
    suffix = filepath.suffix.lower()

    if suffix == ".json":
        return _load_json(filepath)
    elif suffix == ".md":
        return _load_markdown(filepath)
    elif suffix == ".txt":
        return _load_text(filepath)
    elif suffix == ".docx":
        return _load_docx(filepath)
    elif suffix == ".xlsx":
        return _load_xlsx(filepath)
    else:
        return _load_json(filepath)


def _save_to_file(entries: list, filepath: Path):
    """Save entries to a file based on its extension."""
    suffix = filepath.suffix.lower()

    if suffix == ".json":
        _save_json(entries, filepath)
    elif suffix == ".md":
        _save_markdown(entries, filepath)
    elif suffix == ".txt":
        _save_text(entries, filepath)
    elif suffix == ".docx":
        _save_docx(entries, filepath)
    elif suffix == ".xlsx":
        _save_xlsx(entries, filepath)
    else:
        _save_json(entries, filepath)


# --- JSON ---

def _load_json(filepath: Path) -> list:
    try:
        return json.loads(filepath.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []


def _save_json(entries: list, filepath: Path):
    filepath.write_text(
        json.dumps(entries, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# --- Markdown ---

def _load_markdown(filepath: Path) -> list:
    """Parse a markdown file into entries. Expects #### headers as entry titles."""
    import re
    content = filepath.read_text(encoding="utf-8")
    entries = []

    # Pattern: #### Title followed by optional status and STAR content
    entry_pattern = re.compile(
        r'####\s+(.+?)\n'
        r'(?:\*(.+?)\*\n)?'
        r'\n?'
        r'(.*?)(?=\n####|\n---|\Z)',
        re.DOTALL
    )

    for match in entry_pattern.finditer(content):
        title = match.group(1).strip()
        status = (match.group(2) or "In Progress").strip()
        body = match.group(3).strip()

        star = {}
        for component in ["Situation", "Task", "Action", "Result"]:
            pat = re.compile(
                rf'\*\*{component}:\*\*\s*(.+?)(?=\*\*(?:Situation|Task|Action|Result):\*\*|\Z)',
                re.DOTALL
            )
            m = pat.search(body)
            if m:
                star[component.lower()] = m.group(1).strip()

        from datetime import datetime
        entries.append({
            "id": title.lower().replace(" ", "_")[:40],
            "title": title,
            "status": status,
            "date_completed": "",
            "situation": star.get("situation", ""),
            "task": star.get("task", ""),
            "actions": [a.strip() for a in star.get("action", "").split("\n") if a.strip()] if star.get("action") else [],
            "result": star.get("result", ""),
            "created_at": datetime.now().isoformat(),
        })

    return entries


def _save_markdown(entries: list, filepath: Path):
    """Save entries as a markdown file with #### headers."""
    lines = ["# Career Tracker\n", ""]

    for entry in entries:
        lines.append(f"#### {entry.get('title', 'Untitled')}")
        status = entry.get("status", "In Progress")
        lines.append(f"*{status}*\n")

        if entry.get("situation"):
            lines.append(f"**Situation:** {entry['situation']}\n")
        if entry.get("task"):
            lines.append(f"**Task:** {entry['task']}\n")
        if entry.get("actions"):
            actions_text = "\n".join(f"- {a}" for a in entry["actions"])
            lines.append(f"**Action:**\n{actions_text}\n")
        if entry.get("result"):
            lines.append(f"**Result:** {entry['result']}\n")

        lines.append("---\n")

    filepath.write_text("\n".join(lines), encoding="utf-8")


# --- Text ---

def _load_text(filepath: Path) -> list:
    """Load from plain text — one entry per non-empty line."""
    from importer import parse_pasted_list
    from datetime import datetime

    content = filepath.read_text(encoding="utf-8")
    items = parse_pasted_list(content)

    entries = []
    for item in items:
        entries.append({
            "id": item.lower().replace(" ", "_")[:40],
            "title": item,
            "status": "In Progress",
            "date_completed": "",
            "situation": "",
            "task": "",
            "actions": [],
            "result": "",
            "created_at": datetime.now().isoformat(),
        })
    return entries


def _save_text(entries: list, filepath: Path):
    """Save as plain text — one title per line."""
    lines = [entry.get("title", "") for entry in entries if entry.get("title")]
    filepath.write_text("\n".join(lines), encoding="utf-8")


# --- DOCX ---

def _load_docx(filepath: Path) -> list:
    """Load entries from a Word document (paragraphs as items)."""
    from importer import import_from_docx
    from datetime import datetime

    items = import_from_docx(str(filepath))
    entries = []
    for item in items:
        entries.append({
            "id": item.lower().replace(" ", "_")[:40],
            "title": item,
            "status": "In Progress",
            "date_completed": "",
            "situation": "",
            "task": "",
            "actions": [],
            "result": "",
            "created_at": datetime.now().isoformat(),
        })
    return entries


def _save_docx(entries: list, filepath: Path):
    """Save entries as a Word document."""
    from exporter import export_to_docx
    export_to_docx(entries, str(filepath))


# --- XLSX ---

def _load_xlsx(filepath: Path) -> list:
    """Load entries from an Excel file."""
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    entries = []
    headers = []

    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri == 0:
            headers = [str(h).lower().strip() if h else "" for h in row]
            continue

        if not any(row):
            continue

        entry = {
            "id": "",
            "title": "",
            "status": "In Progress",
            "date_completed": "",
            "situation": "",
            "task": "",
            "actions": [],
            "result": "",
            "created_at": datetime.now().isoformat(),
        }

        for ci, value in enumerate(row):
            if ci >= len(headers) or not value:
                continue
            col = headers[ci]
            val = str(value).strip()

            if col in ("title", "name", "project", "accomplishment", "item"):
                entry["title"] = val
            elif col in ("status", "state"):
                entry["status"] = val
            elif col in ("date", "date_completed", "completed"):
                entry["date_completed"] = val
            elif col in ("situation", "context", "problem"):
                entry["situation"] = val
            elif col in ("task", "responsibility"):
                entry["task"] = val
            elif col in ("action", "actions", "what i did", "steps"):
                entry["actions"] = [a.strip() for a in val.split(";") if a.strip()]
            elif col in ("result", "results", "outcome", "impact"):
                entry["result"] = val

        if entry["title"]:
            entry["id"] = entry["title"].lower().replace(" ", "_")[:40]
            entries.append(entry)

    wb.close()
    return entries


def _save_xlsx(entries: list, filepath: Path):
    """Save entries as an Excel file."""
    from exporter import export_to_xlsx
    export_to_xlsx(entries, str(filepath))
