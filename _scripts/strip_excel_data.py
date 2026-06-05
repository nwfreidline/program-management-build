"""
Strip all data rows from Excel files, keeping only header rows.
Preserves formatting, column widths, and sheet structure.

Usage: python _scripts/strip_excel_data.py
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


TEMPLATES_DIR = Path(__file__).parent.parent / "05-Maintenance-Scheduling" / "Schedule-Templates"


def strip_workbook(filepath: Path):
    """Remove all data rows (keep row 1 headers) from all sheets."""
    wb = load_workbook(filepath)
    sheets_stripped = []

    for ws in wb.worksheets:
        max_row = ws.max_row
        if max_row <= 1:
            sheets_stripped.append(f"  {ws.title}: already empty (header only)")
            continue

        # Delete all rows after header (row 1)
        ws.delete_rows(2, max_row - 1)
        sheets_stripped.append(f"  {ws.title}: removed {max_row - 1} data rows")

    wb.save(filepath)
    return sheets_stripped


def main():
    xlsx_files = sorted(TEMPLATES_DIR.glob("*.xlsx"))

    if not xlsx_files:
        print("No .xlsx files found in Schedule-Templates/")
        return

    print(f"Stripping data from {len(xlsx_files)} Excel files...\n")

    for filepath in xlsx_files:
        print(f"📄 {filepath.name}")
        try:
            results = strip_workbook(filepath)
            for r in results:
                print(r)
        except Exception as e:
            print(f"  ✗ Error: {e}")
        print()

    print("✓ Done. All files now contain header rows only.")


if __name__ == "__main__":
    main()
